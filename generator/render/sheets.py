"""Tabular rendering: model-derived tables -> .xlsx (openpyxl) and .csv.

A *table* is ``{"title", "headers": [str], "rows": [[...]], "currency_cols": [int]}``.

Determinism: .xlsx is a zip; we pin the workbook core dates and re-pack the archive
via the shared normalizer, so the same content -> byte-identical .xlsx. .csv is plain
UTF-8 text (already deterministic). Both carry a visible synthetic marker on row 1.
"""

from __future__ import annotations

import csv as _csv
import io
import re
from pathlib import Path

from .. import COMPANY_NAME, SYNTHETIC_MARKER
from ._repro import FIXED_DT, normalize_zip

_CURRENCY_FMT = "$#,##0.00"

# openpyxl stamps dcterms:modified with the wall clock at save(); pin it back to
# the fixed anchor so the .xlsx is byte-reproducible.
_FIXED_W3CDTF = FIXED_DT.strftime("%Y-%m-%dT%H:%M:%SZ").encode()
_MODIFIED_RE = re.compile(rb"(<dcterms:modified[^>]*>)[^<]*(</dcterms:modified>)")


def _pin_modified(name: str, data: bytes) -> bytes:
    if name == "docProps/core.xml":
        return _MODIFIED_RE.sub(rb"\g<1>" + _FIXED_W3CDTF + rb"\g<2>", data)
    return data


def write_xlsx(tables: list[dict], out_path: Path) -> Path:
    """Render one or more tables (one sheet each) to a reproducible .xlsx."""
    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    marker_font = Font(bold=True, color="C0392B")
    header_font = Font(bold=True)
    for i, table in enumerate(tables):
        ws = wb.active if i == 0 else wb.create_sheet()
        ws.title = table["title"][:31]  # Excel sheet-name limit
        ws.append([SYNTHETIC_MARKER])
        ws["A1"].font = marker_font
        ws.append(table["headers"])
        for col in range(1, len(table["headers"]) + 1):
            ws.cell(row=2, column=col).font = header_font
        currency_cols = set(table.get("currency_cols", []))
        for row in table["rows"]:
            ws.append(row)
            r = ws.max_row
            for c in currency_cols:
                ws.cell(row=r, column=c + 1).number_format = _CURRENCY_FMT

    props = wb.properties
    props.creator = COMPANY_NAME
    props.created = FIXED_DT
    props.modified = FIXED_DT
    props.subject = SYNTHETIC_MARKER
    props.title = tables[0]["title"] if tables else COMPANY_NAME

    bio = io.BytesIO()
    wb.save(bio)
    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_bytes(normalize_zip(bio.getvalue(), rewrite=_pin_modified))
    return out_path


def write_csv(table: dict, out_path: Path) -> Path:
    """Render a single table to UTF-8 CSV (marker row, then headers, then rows)."""
    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    buf = io.StringIO()
    writer = _csv.writer(buf, lineterminator="\n")
    writer.writerow([SYNTHETIC_MARKER])
    writer.writerow(table["headers"])
    for row in table["rows"]:
        writer.writerow(row)
    out_path.write_bytes(buf.getvalue().encode("utf-8"))
    return out_path
